import openpyxl as xl
import openpyxl.styles as xs
from openpyxl.utils import get_column_letter as gcl
from datetime import datetime, timedelta


class Scraper:
    def __init__(self, wko_sheet_path, logs=False):
        if (logs): print("Starting...")
        self.wb = xl.load_workbook(wko_sheet_path, data_only=True)
        if (logs): print("WKO Loaded")
        self.ws = self.wb.active
        self.cols = self.ws.max_column
        if (logs): print("Processing...")
        self.get_data()
        if (logs): print("Data Processed")


    def add_days(self, str_time, wko_of_week):
        days_to_add = {2: 1, 3: 3, 4: 4}.get(wko_of_week, 0)
        dt = datetime.strptime(str_time, "%d/%m/%y")
        dt += timedelta(days=days_to_add)
        return dt.strftime("%d/%m/%y")

    def get_data(self):
        self.wko_data = []           # List[Dict[col_num, date, exc]]
        self.wko_data_attended = []  # List[Dict[col_num, date, exc]]
        self.exercise_dict = {}      # Dict[]

        for col_idx in range(self.cols):
            header_val = str(self.ws[gcl(col_idx + 1) + "1"].value)

            if header_val != "None":
                # Try parsing either short or full datetime formats
                if header_val[2] == "-" or header_val[3] == "-":
                    date_str = header_val.split(" ")[2]
                else:
                    dt = datetime.strptime(header_val, "%Y-%m-%d %H:%M:%S")
                    date_str = dt.strftime("%d/%m/%y")

                col_num = col_idx + 1
                temp = []
                wko_of_week = 0

                for row in range(2, 30):
                    exc = str(self.ws[gcl(col_num) + str(row)].value)
                    rp1 = str(self.ws[gcl(col_num + 1) + str(row)].value)
                    rp2 = str(self.ws[gcl(col_num + 2) + str(row)].value)
                    rp3 = str(self.ws[gcl(col_num + 3) + str(row)].value)

                    # Convert 'None' to empty string
                    rp1 = "" if rp1 == "None" else rp1
                    rp2 = "" if rp2 == "None" else rp2
                    rp3 = "" if rp3 == "None" else rp3

                    if exc == "None":
                        if temp:
                            wko_of_week += 1
                            date_with_offset = self.add_days(date_str, wko_of_week)

                            wko_entry = {
                                "col_num": col_num,
                                "date": date_with_offset,
                                "exc": temp
                            }
                            self.wko_data.append(wko_entry)

                            for ex in wko_entry["exc"]:
                                parts = ex[0].split()
                                try:
                                    if parts[-1].endswith("kg"):
                                        weight = float(parts[-1][:-2])
                                        exercise = " ".join(parts[:-1])
                                    elif parts[-1].endswith("lb"):
                                        weight = float(parts[-1][:-2]) / 2
                                        exercise = " ".join(parts[:-1])
                                    else: raise ValueError()
                                except:
                                    weight = -1
                                    exercise = ex[0]

                                try:
                                    weight = 50 if weight == -1 else weight
                                    score = round(weight * (1 + float(ex[1]) / 30), 2)
                                    data = [wko_entry["date"], score, weight, (float(ex[1]), float(ex[2]), float(ex[3]))]
                                except: continue

                                self.exercise_dict.setdefault(exercise, []).append(data)

                            if any(ex[1] != "" for ex in temp):
                                self.wko_data_attended.append(wko_entry)

                        temp = []
                    else:
                        temp.append([exc, rp1, rp2, rp3])


class ExcelTools:
    @staticmethod
    def set_col_widths(ws, widths, start_col=1):
        """Set column widths starting at a specific column."""
        for i, width in enumerate(widths):
            col_letter = gcl(start_col + i)
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def apply_border(ws, col, max_row=200):
        """Apply a thick left border to a range of rows in a column."""
        border = xs.Side(border_style="thick", color="000000")
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = xs.Border(left=border)

    @staticmethod
    def highlight_cell_range(ws, col_start, row, font):
        """Apply the same font formatting across 4 columns in a single row."""
        for i in range(4):
            cell = ws.cell(row=row, column=col_start + i)
            cell.font = font

    @staticmethod
    def center_align_range(ws, min_row=1, max_row=100, min_col=1, max_col=26):
        """Center-align a range of cells."""
        align = xs.Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = align

    @staticmethod
    def merge_and_style_header(ws, start_col, end_col, row, value, bold=True, size=14):
        """Merge a header row and apply center alignment and font styling."""
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col, value=value)
        cell.font = xs.Font(bold=bold, size=size)
        cell.alignment = xs.Alignment(horizontal="center", vertical="center")

    @staticmethod
    def get_highlight_font(date, attended_dates):
        """Return a font object if the date matches any of the highlight conditions."""
        highlight_map = {
            -1: ("0070C0", True),
            -11: ("963634", False),
            -3: ("782170", False)
        }

        for offset, (color, bold) in highlight_map.items():
            dates = [attended_dates[offset]["date"], attended_dates[offset - 1]["date"]]
            if date in dates:
                return xs.Font(color=color, bold=bold, italic=not bold)
        return None