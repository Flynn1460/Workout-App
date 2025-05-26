import openpyxl as xl, openpyxl.styles as xs
import sheet_parser
from openpyxl.utils import get_column_letter as gcl

LOGS = True

scrape_obj = sheet_parser.Scraper("input.xlsx", logs=LOGS)
tools      = sheet_parser.ExcelTools()

wko_data          = scrape_obj.wko_data
wko_data_attended = scrape_obj.wko_data_attended
exercise_dict     = scrape_obj.exercise_dict

write_wb = xl.Workbook()
write_ws = write_wb.active


def create_workout_log(ws):
    ws["A1"].value = "Workout Date"
    ws["B1"].value = "Exercises"

    ws["A1"].font = xs.Font(bold=True)
    ws["B1"].font = xs.Font(bold=True)

    tools.set_col_widths(ws, [12.5, 23.5, 4, 4, 4], start_col=1)

    for i, workout in enumerate(wko_data):
        ws[f"A{i + 2}"].value = workout["date"]
        for j, ex in enumerate(workout["exc"]):
            for k in range(4):
                ws[f"{gcl(2 + (4 * j) + k)}{i + 2}"].value = ex[k]

def create_filtered_exercises(ws):
    starting_col = 1

    for exercise, entries in sorted(exercise_dict.items(), key=lambda x: len(x[1]), reverse=True):
        tools.merge_and_style_header(ws, start_col=starting_col, end_col=starting_col+3, row=1, value=exercise)

        tools.set_col_widths(ws, [10, 5, 3.5, 3.5], start_col=starting_col)
        tools.apply_border(ws, col=starting_col, max_row=200)

        entries.sort(key=lambda x: x[1], reverse=True)

        for row_idx, entry in enumerate(entries):
            date, score, wt, reps = entry
            row_num = row_idx + 2

            for offset, val in enumerate([date, score, wt, reps[0]]):
                ws[f"{gcl(starting_col + offset)}{row_num}"] = val

            font = tools.get_highlight_font(date, wko_data_attended)
            if font:
                tools.highlight_cell_range(ws, starting_col, row_num, font)

        starting_col += 4

def create_stats(ws):
    def write_col(start_col, data, title):
        total = len(data)
        attended = sum(1 for x in data if any(e[1] for e in x["exc"]))
        perc = round((attended / total) * 100, 2)

        tools.set_col_widths(ws, [20, 6.9], start_col=start_col)
        tools.merge_and_style_header(ws, start_col=start_col, end_col=start_col + 1, row=1, value=title)

        labels = ["Total Workouts", "Attended Workouts", "Attended Workout Perc"]
        values = [total, attended, f"{perc}%"]

        for i in range(3):
            ws[f"{gcl(start_col)}{i + 2}"].value = labels[i]
            ws[f"{gcl(start_col + 1)}{i + 2}"].value = values[i]

    write_col(1, wko_data, "Total")
    write_col(3, wko_data[-48:], "3 Months")
    write_col(5, wko_data[-8:], "2 Weeks")

    tools.center_align_range(ws, min_row=1, max_row=100, min_col=1, max_col=26)


write_ws.title = "FILTERED EXERCISES"
create_filtered_exercises(write_ws)

write_ws = write_wb.create_sheet("STATS")
create_stats(write_ws)

write_ws = write_wb.create_sheet("WORKOUT LOG")
create_workout_log(write_ws)

write_wb.save("output.xlsx")
